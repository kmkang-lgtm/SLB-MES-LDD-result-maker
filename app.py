import io
import os
import zipfile
import tempfile
from datetime import datetime

import streamlit as st
import openpyxl
import pandas as pd

import engine
from summary_engine import build_from_zip_bytes
from dashboard_engine import (
    build_dashboard_from_zip_bytes,
    build_dashboard_from_file_bytes,
)

from ui_error import show_error, run_with_ui_error
from validator import pre_validate


# ---------------------------
# App Config / Password Gate
# ---------------------------
st.set_page_config(page_title="SLB MES Result Generator", layout="wide")

APP_TITLE = "SLB MES Result Generator"
st.title(APP_TITLE)

if "APP_PASSWORD" not in st.secrets:
    st.error("APP_PASSWORD가 설정되지 않았습니다. Streamlit secrets를 확인하세요.")
    st.stop()

if "authed" not in st.session_state:
    st.session_state.authed = False

with st.sidebar:
    st.header("접속")
    if not st.session_state.authed:
        pw = st.text_input("비밀번호", type="password")
        if st.button("로그인", use_container_width=True):
            if pw == st.secrets["APP_PASSWORD"]:
                st.session_state.authed = True
                st.success("로그인 성공")
            else:
                st.error("비밀번호가 올바르지 않습니다.")
    else:
        st.success("인증됨")
        if st.button("로그아웃", use_container_width=True):
            st.session_state.authed = False
            st.rerun()

if not st.session_state.authed:
    st.stop()


# ---------------------------
# Helpers
# ---------------------------
def _now_mmdd() -> str:
    return datetime.now().strftime("%m.%d")


def _zip_bytes_from_folder(folder_path: str, zip_name: str) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for root, _, files in os.walk(folder_path):
            for f in files:
                full = os.path.join(root, f)
                rel = os.path.relpath(full, folder_path)
                zf.write(full, rel)
    return buf.getvalue()


def _collect_input_files(uploaded_files, uploaded_zip, input_source: str | None) -> list[tuple[str, bytes]]:
    """
    return: [(filename, file_bytes), ...]
    input_source:
      - "ZIP 사용" / "엑셀 파일 사용" / None
    """
    items: list[tuple[str, bytes]] = []

    if input_source == "ZIP 사용":
        if uploaded_zip is None:
            return []
        z = zipfile.ZipFile(io.BytesIO(uploaded_zip.getvalue()))
        for n in z.namelist():
            if n.lower().endswith(".xlsx") and not n.startswith("__MACOSX/"):
                items.append((os.path.basename(n), z.read(n)))
        return items

    if input_source == "엑셀 파일 사용":
        if not uploaded_files:
            return []
        for uf in uploaded_files:
            items.append((uf.name, uf.getvalue()))
        return items

    return []


def _try_copy_default(src: str, dst: str) -> None:
    try:
        if os.path.exists(src):
            with open(src, "rb") as fsrc, open(dst, "wb") as fdst:
                fdst.write(fsrc.read())
    except Exception:
        pass


def _prepare_templates_on_temp(
    default_khd: str,
    default_wph: str,
    tpl_khd_upload,
    tpl_wph_upload,
    tmp_root_prefix: str,
) -> dict[str, str]:
    tmp_root = tempfile.mkdtemp(prefix=tmp_root_prefix)
    tpl_dir = os.path.join(tmp_root, "templates")
    os.makedirs(tpl_dir, exist_ok=True)

    khd_path = os.path.join(tpl_dir, "TEMPLATE_KHD.xlsx")
    wph_path = os.path.join(tpl_dir, "TEMPLATE_WPH.xlsx")

    _try_copy_default(default_khd, khd_path)
    _try_copy_default(default_wph, wph_path)

    if tpl_khd_upload is not None:
        with open(khd_path, "wb") as f:
            f.write(tpl_khd_upload.getvalue())

    if tpl_wph_upload is not None:
        with open(wph_path, "wb") as f:
            f.write(tpl_wph_upload.getvalue())

    return {"KHD": khd_path, "WPH": wph_path}


def _get_template_sheetnames(templates: dict[str, str]) -> dict[str, list[str]]:
    out: dict[str, list[str]] = {}
    for dtype, path in templates.items():
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        out[dtype] = wb.sheetnames
        wb.close()
    return out


def _parse_selected_hours(selected_labels: list[int]) -> list[int]:
    # UI label 1..24 -> engine hour 0..23
    out = []
    for v in selected_labels:
        out.append(0 if v == 24 else int(v))
    return sorted(set(out))


def _hours_to_labels(hours: list[int]) -> list[int]:
    return [24 if h == 0 else h for h in hours]


def _apply_recommend_exclude(exclude_labels: list[int]) -> None:
    """
    Streamlit widget(key='hour_filter') 업데이트는 on_click 콜백에서 수행해야 안정적임.
    """
    cur = set(st.session_state.get("hour_filter", []))
    new = sorted(list(cur - set(exclude_labels)))
    # 전부 빠져버리면 적용하지 않음(사용자 실수 방지)
    if new:
        st.session_state["hour_filter"] = new


# ---------------------------
# Session State defaults
# ---------------------------
st.session_state.setdefault("zip_bytes", None)
st.session_state.setdefault("zip_filename", None)

st.session_state.setdefault("validation_result", None)
st.session_state.setdefault("validation_ok", False)

# 시간 multiselect는 key 기반으로 제어
hour_labels = list(range(1, 25))
st.session_state.setdefault("hour_filter", hour_labels[:])  # 최초 기본은 전체

# 생성 범위 기본
st.session_state.setdefault("include_lanes", ["1Lane", "2Lane"])
st.session_state.setdefault("include_dtypes", ["KHD", "WPH"])

# 다운로드 형태 기본
st.session_state.setdefault("dl_zip", True)
st.session_state.setdefault("dl_each", False)


# ---------------------------
# Section 1: Result 생성
# ---------------------------
st.subheader("1) Result 생성 (원본 → Lane별 Result 엑셀 → ZIP/개별 다운로드)")

colA, colB = st.columns([2, 1], gap="large")

with colA:
    st.markdown("### 입력")
    uploaded_files = st.file_uploader(
        "원본 MES 엑셀(.xlsx) 여러 개 업로드",
        type=["xlsx"],
        accept_multiple_files=True,
        key="uploader_xlsx_multi",
    )
    uploaded_zip = st.file_uploader(
        "또는 원본 폴더를 ZIP으로 업로드 (.zip 안에 .xlsx 포함)",
        type=["zip"],
        accept_multiple_files=False,
        key="uploader_xlsx_zip",
    )

    # ✅ 둘 다 업로드 시 선택 / 하나만 있으면 자동
    input_source = None
    if uploaded_files and uploaded_zip:
        input_source = st.radio(
            "입력 소스 선택",
            ["ZIP 사용", "엑셀 파일 사용"],
            horizontal=True,
            key="input_source_choice",
        )
    elif uploaded_zip:
        input_source = "ZIP 사용"
    elif uploaded_files:
        input_source = "엑셀 파일 사용"

with colB:
    default_khd = "templates/TEMPLATE_KHD.xlsx"
    default_wph = "templates/TEMPLATE_WPH.xlsx"

    # ✅ 옵션/템플릿/시간필터 숨김
    with st.expander("⚙️ 고급 옵션(템플릿/시간/Raw)", expanded=False):
        st.markdown("### 템플릿")
        tpl_khd_upload = st.file_uploader("KHD 템플릿 업로드(선택)", type=["xlsx"], key="tpl_khd")
        tpl_wph_upload = st.file_uploader("WPH 템플릿 업로드(선택)", type=["xlsx"], key="tpl_wph")

        st.markdown("### 옵션")
        raw_end_row = st.number_input(
            "Raw 끝행(raw_end_row)",
            min_value=20,
            max_value=5000,
            value=100,
            step=10,
            help="템플릿 차트가 참조하는 Raw 데이터 영역의 마지막 행",
            key="raw_end_row",
        )

        st.multiselect(
            "시간 필터(선택한 시간대만 반영)",
            options=hour_labels,
            key="hour_filter",
            help="24는 자정(00시)로 처리됩니다.",
        )

    # ✅ 생성 범위 선택(보이게)
    st.markdown("### 생성 범위 선택")
    st.multiselect(
        "Lane 선택",
        ["1Lane", "2Lane"],
        default=st.session_state["include_lanes"],
        key="include_lanes",
    )
    st.multiselect(
        "Type 선택",
        ["KHD", "WPH"],
        default=st.session_state["include_dtypes"],
        key="include_dtypes",
    )

    # ✅ 다운로드 형태 선택(보이게)
    st.markdown("### 다운로드 형태")
    st.checkbox("ZIP 다운로드", value=st.session_state["dl_zip"], key="dl_zip")
    st.checkbox("개별 엑셀 다운로드", value=st.session_state["dl_each"], key="dl_each")

# hour 필터 파싱
selected_labels = st.session_state["hour_filter"]
selected_hours = _parse_selected_hours(selected_labels)
raw_end_row_val = int(st.session_state.get("raw_end_row", 100))


# ---------------------------
# Validation + 4개 미리보기 + 추천 적용
# ---------------------------
st.markdown("### 사전 점검 + 미리보기 (ZIP 만들기 전)")

validate_btn = st.button("🔍 사전 점검 실행", use_container_width=True, key="btn_validate")

if validate_btn:
    try:
        inputs = _collect_input_files(uploaded_files, uploaded_zip, input_source)
        if not inputs:
            st.warning("원본 엑셀(.xlsx) 파일을 업로드하거나 ZIP을 업로드하세요.")
            st.stop()

        templates_for_validation = _prepare_templates_on_temp(
            default_khd=default_khd,
            default_wph=default_wph,
            tpl_khd_upload=st.session_state.get("tpl_khd"),
            tpl_wph_upload=st.session_state.get("tpl_wph"),
            tmp_root_prefix="mes_validate_",
        )
        # 위에서 uploader를 expander 안에서 만들었기 때문에
        # 직접 변수(tpl_khd_upload/tpl_wph_upload)를 쓰는 쪽이 더 안전함:
        # -> 아래 두 줄로 교체
        # templates_for_validation = _prepare_templates_on_temp(default_khd, default_wph, tpl_khd_upload, tpl_wph_upload, "mes_validate_")

        template_sheetnames = _get_template_sheetnames(templates_for_validation)

        with st.spinner("사전 점검 중..."):
            vr = pre_validate(
                input_files=inputs,
                template_sheetnames=template_sheetnames,
                selected_hours=selected_hours,
                low_count_threshold=3,
            )

        st.session_state["validation_result"] = vr
        st.session_state["validation_ok"] = bool(vr.get("ok", False))

    except Exception as e:
        st.session_state["validation_result"] = None
        st.session_state["validation_ok"] = False
        show_error(e)

vr = st.session_state.get("validation_result")
if vr:
    if vr.get("ok"):
        st.success("사전 점검 통과! (치명 에러 없음)")
    else:
        st.error("사전 점검 실패: 에러를 해결해야 Result를 생성할 수 있습니다.")

    # 치명 에러
    if vr.get("errors"):
        st.markdown("#### ❌ 에러(해결 필요)")
        for e in vr["errors"]:
            show_error(e)

    # 참고 경고
    if vr.get("warnings"):
        st.markdown("#### ⚠️ 경고(그래프/데이터 이상 가능성)")
        for w in vr["warnings"][:200]:
            st.warning(w)

    # 전역 추천 제외 시간(공백 시간 기준)
    rec = vr.get("recommend_global", {})
    exclude_labels = rec.get("exclude_labels", [])

    cols = st.columns([2, 1])
    with cols[0]:
        st.info(
            "추천 제외 시간(1시간 버킷 데이터 0개 기준): "
            + (", ".join(f"{h:02d}" for h in exclude_labels) if exclude_labels else "없음")
        )
    with cols[1]:
        st.button(
            "✅ 추천 시간 제외 적용(원클릭)",
            use_container_width=True,
            key="btn_apply_reco",
            on_click=_apply_recommend_exclude,
            args=(exclude_labels,),
        )

    # 파일별 4개 미리보기 (KHD/WPH x 1Lane/2Lane)
    st.markdown("#### 👀 4개 결과 미리보기 (대표 item 1개씩)")

    by_file = vr.get("by_file", {})
    for fname, pack in by_file.items():
        with st.expander(f"📄 {fname} 미리보기", expanded=True):
            previews = pack.get("previews", {})

            order = [("KHD", "1Lane"), ("KHD", "2Lane"), ("WPH", "1Lane"), ("WPH", "2Lane")]
            c1, c2 = st.columns(2, gap="large")
            slot_cols = [c1, c2, c1, c2]

            for idx, key in enumerate(order):
                dtype, lane = key
                col = slot_cols[idx]
                with col:
                    data = previews.get((dtype, lane))
                    if not data:
                        st.warning(f"{dtype} {lane}: 데이터 없음(또는 dtype 감지 실패)")
                        continue

                    st.markdown(f"**{dtype} {lane}**")
                    st.caption(f"대표 item: {data['item']}")
                    st.caption(f"기간: {data['date_range']} / parse OK: {data['parse_ok']:.0%}")

                    miss = data.get("missing_hours", [])
                    low = data.get("low_hours", [])

                    if miss:
                        st.error(
                            f"1시간 공백(데이터 0개): {', '.join(f'{h:02d}' for h in _hours_to_labels(miss))}"
                        )
                    else:
                        st.success("1시간 공백 없음")

                    if low:
                        st.warning(
                            f"샘플 수 부족(<3): {', '.join(f'{h:02d}' for h in _hours_to_labels(low))}"
                        )

                    s = pd.Series(data["hourly_avg_series"])
                    s.index = [f"{i:02d}" for i in range(1, 25)]
                    st.line_chart(s)

            if vr.get("summary_rows"):
                st.markdown("##### 📋 요약 테이블")
                rows = [r for r in vr["summary_rows"] if r.get("File") == fname]
                if rows:
                    st.dataframe(rows, use_container_width=True)


# ---------------------------
# Result 생성 버튼(Validation 통과 시 활성)
# ---------------------------
make_btn = st.button(
    "Result 생성하기",
    use_container_width=True,
    disabled=not st.session_state.get("validation_ok", False),
)

if make_btn:
    if not st.session_state.get("validation_ok", False):
        st.warning("먼저 사전 점검을 실행하고, 통과한 뒤 Result를 생성하세요.")
        st.stop()

    inputs = _collect_input_files(uploaded_files, uploaded_zip, input_source)
    if not inputs:
        st.warning("원본 엑셀(.xlsx) 파일을 업로드하거나 ZIP을 업로드하세요.")
        st.stop()

    # 템플릿 준비
    # (expander 안에서 만든 uploader 변수는 scope상 안전하므로 직접 쓰는 것이 가장 확실)
    try:
        tpl_khd_upload = st.session_state.get("tpl_khd")
        tpl_wph_upload = st.session_state.get("tpl_wph")
    except Exception:
        tpl_khd_upload = None
        tpl_wph_upload = None

    templates = _prepare_templates_on_temp(
        default_khd=default_khd,
        default_wph=default_wph,
        tpl_khd_upload=tpl_khd_upload,
        tpl_wph_upload=tpl_wph_upload,
        tmp_root_prefix="mes_run_tpl_",
    )

    tmp_root = tempfile.mkdtemp(prefix="mes_run_")
    out_dir = os.path.join(tmp_root, "outputs")
    os.makedirs(out_dir, exist_ok=True)

    include_lanes = st.session_state.get("include_lanes", ["1Lane", "2Lane"])
    include_dtypes = st.session_state.get("include_dtypes", ["KHD", "WPH"])

    def _run_make():
        created_all: list[str] = []
        for fname, fbytes in inputs:
            in_path = os.path.join(tmp_root, "inputs", fname)
            os.makedirs(os.path.dirname(in_path), exist_ok=True)
            with open(in_path, "wb") as f:
                f.write(fbytes)

            created = engine.make_results_for_input(
                input_path=in_path,
                templates=templates,
                output_dir=out_dir,
                raw_end_row=int(raw_end_row_val),
                selected_hours=_parse_selected_hours(st.session_state["hour_filter"]),
                include_lanes=include_lanes,
                include_dtypes=include_dtypes,
            )
            created_all.extend(created)

        if not created_all:
            raise Exception("생성된 결과 파일이 없습니다. (선택한 Lane/Type 범위를 확인하세요.)")
        return created_all

    created_files = run_with_ui_error(_run_make, spinner_text="Result 생성 중...")
    if created_files is None:
        st.stop()

    # ✅ ZIP 다운로드(선택)
    if st.session_state.get("dl_zip", True):
        zip_name = f"SLB_MES_Result_Package_{_now_mmdd()}.zip"
        zip_bytes = _zip_bytes_from_folder(out_dir, zip_name)

        st.session_state["zip_bytes"] = zip_bytes
        st.session_state["zip_filename"] = zip_name

        st.success("Result ZIP 생성 완료!")
        st.download_button(
            "⬇️ Result ZIP 다운로드",
            data=zip_bytes,
            file_name=zip_name,
            mime="application/zip",
            use_container_width=True,
            key="dl-result-zip",
        )

    # ✅ 개별 엑셀 다운로드(선택)
    if st.session_state.get("dl_each", False):
        st.markdown("### 개별 Result 엑셀 다운로드")
        for p in created_files:
            bn = os.path.basename(p)
            with open(p, "rb") as f:
                st.download_button(
                    f"⬇️ {bn}",
                    data=f.read(),
                    file_name=bn,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"dl_each_{bn}",
                )

    with st.expander("생성된 파일 목록"):
        for p in created_files:
            st.write("-", os.path.basename(p))


# ---------------------------
# Section 2: Deviation Summary 생성
# ---------------------------
st.divider()
st.subheader("2) Deviation Summary 생성 (Result ZIP → Summary 엑셀)")

has_last_zip = bool(st.session_state.get("zip_bytes"))

use_last = st.checkbox(
    "바로 이전에 생성한 Result ZIP으로 Summary 만들기(업로드 없이)",
    value=has_last_zip,
    disabled=not has_last_zip,
)

zip_upload_for_summary = None
if not use_last:
    zip_upload_for_summary = st.file_uploader(
        "기존 SLB_MES_Result_Package_XX.XX.zip 업로드",
        type=["zip"],
        key="uploader_summary_zip",
    )
else:
    st.info(f"직전 결과 사용: {st.session_state.get('zip_filename', 'results.zip')}")

if st.button("📌 Summary 생성하기", use_container_width=True, key="btn_summary"):
    try:
        if use_last:
            zip_bytes = st.session_state["zip_bytes"]
            zip_name = st.session_state.get("zip_filename", "results.zip")
        else:
            if zip_upload_for_summary is None:
                st.warning("Summary를 만들 ZIP 파일을 업로드하세요.")
                st.stop()
            zip_bytes = zip_upload_for_summary.getvalue()
            zip_name = zip_upload_for_summary.name

        with st.spinner("Summary 생성 중..."):
            summary_name, summary_bytes = build_from_zip_bytes(zip_bytes, zip_name)

        st.success("Summary 생성 완료!")
        st.download_button(
            "⬇️ Summary 다운로드",
            data=summary_bytes,
            file_name=summary_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="dl-summary",
        )

    except Exception as e:
        show_error(e)


# ---------------------------
# Section 3: Dashboard 생성
# ---------------------------
st.divider()
st.subheader("3) Dashboard 생성 (Summary 여러 개 → 기간 Dashboard 엑셀)")

dash_mode = st.radio(
    "입력 방식",
    options=["Summary ZIP 업로드", "Summary 파일 여러 개 업로드"],
    horizontal=True,
    key="dash_mode",
)

dash_zip = None
dash_files = None

if dash_mode == "Summary ZIP 업로드":
    dash_zip = st.file_uploader(
        "Summary 엑셀들이 들어있는 ZIP 업로드",
        type=["zip"],
        key="uploader_dash_zip",
    )
else:
    dash_files = st.file_uploader(
        "Summary 엑셀(.xlsx) 여러 개 업로드",
        type=["xlsx"],
        accept_multiple_files=True,
        key="uploader_dash_files",
    )

if st.button("📊 Dashboard 생성하기", use_container_width=True, key="btn_dash"):
    try:
        with st.spinner("Dashboard 생성 중..."):
            if dash_mode == "Summary ZIP 업로드":
                if dash_zip is None:
                    st.warning("Summary ZIP을 업로드하세요.")
                    st.stop()
                dash_name, dash_bytes = build_dashboard_from_zip_bytes(dash_zip.getvalue(), dash_zip.name)
            else:
                if not dash_files:
                    st.warning("Summary 파일을 하나 이상 업로드하세요.")
                    st.stop()
                file_items = [(f.name, f.getvalue()) for f in dash_files]
                dash_name, dash_bytes = build_dashboard_from_file_bytes(file_items)

        st.success("Dashboard 생성 완료!")
        st.download_button(
            "⬇️ Dashboard 다운로드",
            data=dash_bytes,
            file_name=dash_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="dl-dashboard",
        )

    except Exception as e:
        show_error(e)
