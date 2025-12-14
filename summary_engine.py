import pandas as pd
import numpy as np
import zipfile
import tempfile
import re
import shutil
import gc
from pathlib import Path


# ---------------------------
# 0) zip 파일명에서 날짜 추출
# ---------------------------
def extract_date_from_zipname(zip_filename: str):
    name = Path(zip_filename).stem
    m = re.search(r"SLB_MES_Result_Package_(\d{1,2}\.\d{1,2})", name)
    if not m:
        return name.split("_")[-1]
    return m.group(1)


# ---------------------------
# 1) 원본 파일 파싱 규칙
#
# - result xlsx는 "Summary" 시트 제외, 각 position(시트명)별로 Raw 영역 값들을 수집
# - Material:
#   - 시트명에 "-2" 포함 → AL2
#   - 시트명에 "-1" 포함 + 동일 네이밍의 "-2" 시트 존재 → AL1
#   - 그 외 → CU
# - Mean/Std/Min/Max/Range 계산
# ---------------------------
def parse_result_xlsx(file_path: str):
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    records = []

    # ✅ 시트명 목록(set)으로 AL1/AL2 매칭(끝이 -1/-2가 아니어도 동작)
    sheet_titles = [ws.title for ws in wb.worksheets if ws.title.lower() != "summary"]
    sheet_set = set(sheet_titles)

    for ws in wb.worksheets:
        if ws.title.lower() == "summary":
            continue

        # Raw 영역: 기존 로직(7행부터, 2열부터) 유지
        vals = []
        for c in range(2, ws.max_column + 1):
            for r in range(7, ws.max_row + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                try:
                    vals.append(float(v))
                except Exception:
                    continue

        if not vals:
            continue

        arr = np.array(vals, dtype=float)
        position = ws.title

        name = str(position)

        # ✅ 기본은 CU
        material = "CU"

        # ✅ 1) "-2"가 들어가면 무조건 AL2 (기존 규칙 유지)
        if "-2" in name:
            material = "AL2"

        # ✅ 2) "-1"이고, 같은 네이밍의 "-2" 시트가 실제로 있으면 AL1
        elif "-1" in name:
            cand = name.replace("-1", "-2", 1)
            material = "AL1" if cand in sheet_set else "CU"

        mean = float(np.mean(arr))
        std = float(np.std(arr, ddof=1)) if arr.size > 1 else 0.0
        mn = float(np.min(arr))
        mx = float(np.max(arr))
        rg = mx - mn

        records.append(
            {
                "Position": position,
                "Material": material,
                "Mean": mean,
                "Std": std,
                "Min": mn,
                "Max": mx,
                "Range": rg,
            }
        )

    wb.close()

    if not records:
        return pd.DataFrame(columns=["Position", "Material", "Mean", "Std", "Min", "Max", "Range"])

    df = pd.DataFrame(records)

    # ✅ Material 정렬 순서 고정: AL1 → AL2 → CU
    order = pd.Categorical(df["Material"], categories=["AL1", "AL2", "CU"], ordered=True)
    df["_mat_order"] = order
    df = df.sort_values(["_mat_order", "Position"]).drop(columns=["_mat_order"]).reset_index(drop=True)
    return df


# ---------------------------
# 2) 레이아웃용: Lane별 (WPH + KHD) 결합
# ---------------------------
def combine_lane_df(wph_df: pd.DataFrame, khd_df: pd.DataFrame, lane_label: str):
    # WPH/KHD 둘 다 Position 기반 outer merge (기존 구조 유지)
    merged = pd.merge(
        wph_df, khd_df, on=["Position", "Material"], how="outer", suffixes=("_WPH", "_KHD")
    )

    # 컬럼 순서 정리
    cols = ["Position", "Material"]
    for k in ["Mean", "Std", "Min", "Max", "Range"]:
        cols += [f"{k}_WPH", f"{k}_KHD"]

    for c in cols:
        if c not in merged.columns:
            merged[c] = np.nan

    merged = merged[cols]

    # lane 라벨(표시용)
    merged.insert(0, "Lane", lane_label)
    return merged


# ---------------------------
# 3) 최종 레이아웃: Lane1 + Lane2 를 "한 시트에" 배치하기 위한 블록 생성
#    (기존 양식 유지: lane1 블록 + lane2 블록을 한 sheet에 붙임)
# ---------------------------
def build_final_layout(lane1_df: pd.DataFrame, lane2_df: pd.DataFrame):
    # 기존 방식: lane1, lane2 각각 따로 쓰되 한 시트에 블록 형태로 배치
    # 여기서는 실제 엑셀 쓰기 단계에서 배치하므로 DF는 그대로 반환
    return lane1_df, lane2_df


# ---------------------------
# 4) 엑셀 저장(양식 반영)
# ---------------------------
def write_summary_excel(output_path: str, lane1_df: pd.DataFrame, lane2_df: pd.DataFrame, sheet_name="Summary"):
    # 기존 양식 그대로 유지 (pandas + openpyxl 스타일)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # ✅ Lane1 제목 행을 추가하므로, lane1 DF는 한 줄 아래(startrow=3)에 씀
        lane1_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=3)

        wb = writer.book
        ws = writer.sheets[sheet_name]

        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # ---------------------------
        # 스타일 정의 (기존 양식 유지)
        # ---------------------------
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=11, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")

        fill_title = PatternFill("solid", fgColor="1F4E79")
        fill_header = PatternFill("solid", fgColor="2F5597")
        fill_lane = PatternFill("solid", fgColor="D9E1F2")
        fill_alt = PatternFill("solid", fgColor="F7F7F7")

        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ---------------------------
        # 상단 제목
        # ---------------------------
        max_col = lane1_df.shape[1]
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        tcell = ws.cell(row=1, column=1, value="Deviation Summary")
        tcell.font = title_font
        tcell.fill = fill_title
        tcell.alignment = center

        # ---------------------------
        # lane1 제목(병합) ✅ 실제 반영되도록 DF 시작 행과 분리
        # ---------------------------
        title1_row = 2
        ws.merge_cells(start_row=title1_row, start_column=1, end_row=title1_row, end_column=max_col)
        t1 = ws.cell(row=title1_row, column=1, value="(Below) Lane1")
        t1.font = Font(bold=True, size=12)
        t1.fill = fill_lane
        t1.alignment = center

        # ---------------------------
        # 헤더 스타일 (lane1 DF의 헤더는 row=4에 위치)
        # ---------------------------
        header_row = 4
        for c in range(1, max_col + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.font = header_font
            cell.fill = fill_header
            cell.alignment = center
            cell.border = border

        # ---------------------------
        # lane1 데이터 영역 스타일
        # ---------------------------
        start_data_row_1 = header_row + 1
        end_data_row_1 = start_data_row_1 + len(lane1_df) - 1

        for r in range(start_data_row_1, end_data_row_1 + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                cell.alignment = left if c in (1, 2, 3) else center
                if (r - start_data_row_1) % 2 == 1:
                    cell.fill = fill_alt

        # ---------------------------
        # lane2 DF를 lane1 아래에 붙여쓰기
        # ---------------------------
        gap = 3
        startrow2 = end_data_row_1 + gap + 1  # lane2 DF 시작(startrow)
        lane2_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=startrow2)

        # lane2 제목(병합)
        title2_row = startrow2 - 1
        ws.merge_cells(start_row=title2_row, start_column=1, end_row=title2_row, end_column=max_col)
        t2 = ws.cell(row=title2_row, column=1, value="(Below) Lane2")
        t2.font = Font(bold=True, size=12)
        t2.fill = fill_lane
        t2.alignment = center

        # lane2 헤더 (lane2 DF 헤더는 startrow2+1)
        header_row2 = startrow2 + 1
        for c in range(1, max_col + 1):
            cell = ws.cell(row=header_row2, column=c)
            cell.font = header_font
            cell.fill = fill_header
            cell.alignment = center
            cell.border = border

        # lane2 데이터
        start_data_row_2 = header_row2 + 1
        end_data_row_2 = start_data_row_2 + len(lane2_df) - 1

        for r in range(start_data_row_2, end_data_row_2 + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                cell.alignment = left if c in (1, 2, 3) else center
                if (r - start_data_row_2) % 2 == 1:
                    cell.fill = fill_alt

        # ---------------------------
        # 열 너비(기본 고정, 기존 느낌 유지)
        # ---------------------------
        widths = {
            1: 10,   # Lane
            2: 20,   # Position
            3: 10,   # Material
        }
        for c in range(4, max_col + 1):
            widths[c] = 12

        for c, w in widths.items():
            ws.column_dimensions[chr(64 + c)].width = w

        # ---------------------------
        # 숫자 포맷
        # ---------------------------
        # Mean/Std/Min/Max/Range: 소수점 3자리
        for r in range(start_data_row_1, end_data_row_2 + 1):
            for c in range(4, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, (int, float, np.number)):
                    cell.number_format = "0.000"

        # ---------------------------
        # Freeze panes / Grid off
        # ---------------------------
        ws.sheet_view.showGridLines = False
        # ✅ Lane1 제목+헤더가 보이도록 헤더 아래로 고정
        ws.freeze_panes = "A5"


# ---------------------------
# 5) ZIP bytes → Summary 생성 (외부 API)
# ---------------------------
def build_from_zip_bytes(zip_bytes: bytes, zip_name: str):
    date_str = extract_date_from_zipname(zip_name)

    tmpdir = tempfile.mkdtemp(prefix="mes_summary_")
    try:
        zip_path = Path(tmpdir) / "input.zip"
        with open(zip_path, "wb") as f:
            f.write(zip_bytes)

        with zipfile.ZipFile(zip_path, "r") as z:
            z.extractall(tmpdir)

        files = list(Path(tmpdir).glob("*.xlsx"))

        # ✅ 파일명 매칭을 더 튼튼하게(엔진 출력명이 달라도 웬만하면 잡도록)
        def pick(dtype: str, lane: str):
            dtype = dtype.lower()
            lane = lane.lower()

            # lane alias들(1lane / 1_lane / lane1 / 1-lane 등)
            lane_aliases = {
                lane,
                lane.replace("lane", "_lane"),
                lane.replace("lane", "-lane"),
                lane.replace("lane", "lane"),  # no-op
            }
            if lane == "1lane":
                lane_aliases |= {"lane1", "1_lane", "1-lane", "lane_1", "lane-1"}
            if lane == "2lane":
                lane_aliases |= {"lane2", "2_lane", "2-lane", "lane_2", "lane-2"}

            # 1) dtype + lane 모두 포함(강)
            for f in files:
                name = f.name.lower().replace(" ", "")
                if dtype in name:
                    for a in lane_aliases:
                        if a.replace(" ", "") in name:
                            return str(f)

            # 2) 마지막 fallback: dtype만 맞고 lane 숫자(1/2)라도 포함하면
            lane_num = "1" if lane == "1lane" else "2"
            for f in files:
                name = f.name.lower().replace(" ", "")
                if dtype in name and lane_num in name:
                    return str(f)

            return None

        wph_1 = pick("wph", "1lane")
        wph_2 = pick("wph", "2lane")
        khd_1 = pick("khd", "1lane")
        khd_2 = pick("khd", "2lane")

        missing = [k for k, v in {
            "wph_1": wph_1,
            "wph_2": wph_2,
            "khd_1": khd_1,
            "khd_2": khd_2,
        }.items() if v is None]

        if missing:
            raise Exception(
                f"Result 파일이 부족합니다: {missing}\n"
                f"ZIP 안 파일명에 WPH/KHD, 1Lane/2Lane 포함 여부를 확인하세요."
            )

        # 파싱
        df_wph_1 = parse_result_xlsx(wph_1)
        df_wph_2 = parse_result_xlsx(wph_2)
        df_khd_1 = parse_result_xlsx(khd_1)
        df_khd_2 = parse_result_xlsx(khd_2)

        # Lane별 결합
        lane1 = combine_lane_df(df_wph_1, df_khd_1, "Lane1")
        lane2 = combine_lane_df(df_wph_2, df_khd_2, "Lane2")

        lane1_df, lane2_df = build_final_layout(lane1, lane2)

        # 출력
        out_name = f"SLB_MES_Deviation_Summary_{date_str}.xlsx"
        out_path = Path(tmpdir) / out_name

        write_summary_excel(str(out_path), lane1_df, lane2_df, sheet_name="Summary")

        with open(out_path, "rb") as f:
            out_bytes = f.read()

        return out_name, out_bytes

    finally:
        # cleanup
        try:
            shutil.rmtree(tmpdir, ignore_errors=True)
        except Exception:
            pass
        gc.collect()
