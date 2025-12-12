import pandas as pd
import numpy as np
import openpyxl
import os
import re
from datetime import datetime

# =========================
# Batch Result Generator
# =========================
# 이 스크립트는 특정 패턴의 원본 엑셀들을 찾아서,
# 템플릿(그래프/서식 포함)에 데이터만 채워 넣은 Result 엑셀을 생성합니다.
#
# - 입력: INPUT_GLOB 패턴에 맞는 원본 엑셀들
# - 출력: output/ 폴더에 SLB_MES_{dtype}_Result_{lane}.xlsx 생성
#
# 참고: Streamlit 앱(app.py + engine.py) 쪽이 더 발전된 구조이며,
#       이 파일은 로컬 배치/구버전 흐름에 가까운 템플릿입니다.
# =========================


# 🔧 배치 입력 파일 패턴 (현재 폴더 기준)
INPUT_GLOB = "25.11.18_*.xlsx"

# 🔧 시간 컬럼 고정 순서(그래프/템플릿 헤더 순서와 맞춰야 함)
HOUR_ORDER = list(range(8, 24)) + [0] + list(range(1, 8))
HOUR_LABELS = list(range(8, 24)) + [24] + list(range(1, 8))

# 🔧 템플릿 파일명(현재 폴더에 있어야 함)
TEMPLATES = {
    "KHD": "TEMPLATE_KHD.xlsx",   # KHD용 Summary/그래프 양식 템플릿
    "WPH": "TEMPLATE_WPH.xlsx",   # WPH용 Summary/그래프 양식 템플릿
}

# 🔧 Raw 데이터 영역 설정(템플릿에서 차트가 참조하는 Raw 범위 끝행)
RAW_END_ROW = 100

# 템플릿 규칙
HEADER_ROW = 3
RAW_START_ROW = 7
RAW_START_COL = 2


LANE_SHEETS = {
    "1Lane": ["1Lane_frt", "1Lane_rr", "1Lane_frt side", "1Lane_rr side"],
    "2Lane": ["2Lane_frt", "2Lane_rr", "2Lane_frt side", "2Lane_rr side"],
}


def detect_dtype(item_name: str) -> str:
    if "KHD" in item_name:
        return "KHD"
    if "WPH" in item_name:
        return "WPH"
    return "UNKNOWN"


def item_to_sheetname(item_name: str) -> str:
    """
    원본 항목명 → 템플릿 시트명 변환
    """
    name = item_name.replace("버스바 ", "")
    name = name.replace(" KHD AVG ", "-").replace(" WPH AVG ", "-")
    name = name.replace("FRT Side", "FS 1").replace("RR Side", "RS 1")
    return name


def safe_to_datetime(col):
    """
    문자열 파싱 실패가 많으면 엑셀 날짜(일련번호)로 재해석
    """
    dt = pd.to_datetime(col, errors="coerce")
    if dt.isna().mean() > 0.5:
        num = pd.to_numeric(col, errors="coerce")
        dt2 = pd.to_datetime(num, unit="d", origin="1899-12-30", errors="coerce")
        if dt2.notna().sum() > dt.notna().sum():
            dt = dt2
    return dt


def load_lane_raw(xl: pd.ExcelFile, sheets):
    dfs = [xl.parse(s) for s in sheets]
    raw = pd.concat(dfs, ignore_index=True)

    raw["항목명"] = raw["항목명"].astype(str)
    raw["dtype"] = raw["항목명"].apply(detect_dtype)

    raw["측정일시"] = safe_to_datetime(raw["측정일시"])
    raw = raw.dropna(subset=["측정일시"])

    raw["hour"] = raw["측정일시"].dt.hour
    raw["val"] = pd.to_numeric(raw["측정값"], errors="coerce")
    return raw


def compute_hour_lists(df_item: pd.DataFrame):
    hour_lists = {
        h: df_item.loc[df_item["hour"] == h, "val"].dropna().tolist()
        for h in HOUR_ORDER
    }
    mins = [min(hour_lists[h]) if hour_lists[h] else 0 for h in HOUR_ORDER]
    maxs = [max(hour_lists[h]) if hour_lists[h] else 0 for h in HOUR_ORDER]
    avgs = [
        (sum(hour_lists[h]) / len(hour_lists[h])) if hour_lists[h] else np.nan
        for h in HOUR_ORDER
    ]
    return hour_lists, mins, maxs, avgs


def update_summary_lane_title_from_template(out_wb, template_wb, lane_key):
    """
    템플릿 summary!B2 문자열의 맨 앞 숫자(1/2)만 lane에 맞게 교체
    """
    if "summary" not in out_wb.sheetnames or "summary" not in template_wb.sheetnames:
        return
    tpl_b2 = template_wb["summary"]["B2"].value or ""
    lane_no = "1" if lane_key.startswith("1") else "2"
    new_b2 = re.sub(r"^[12]", lane_no, str(tpl_b2))
    out_wb["summary"]["B2"].value = new_b2


def fill_data_into_ws(
    ws,
    dtype,
    sheet_name,
    hour_lists,
    mins,
    maxs,
    avgs,
):
    """
    템플릿 서식/차트는 유지하고 값만 채움
    """
    ws.cell(row=2, column=2).value = f"{dtype} {sheet_name}"

    # header row (시간 라벨)
    for i in range(len(HOUR_ORDER)):
        c = RAW_START_COL + i
        ws.cell(row=HEADER_ROW, column=c).value = HOUR_LABELS[i]

    # min/max/avg rows
    for i in range(len(HOUR_ORDER)):
        c = RAW_START_COL + i
        ws.cell(row=4, column=c).value = mins[i]
        ws.cell(row=5, column=c).value = maxs[i]
        ws.cell(row=6, column=c).value = avgs[i]

    # raw 영역 클리어
    for r in range(RAW_START_ROW, RAW_END_ROW + 1):
        for i in range(len(HOUR_ORDER)):
            c = RAW_START_COL + i
            ws.cell(row=r, column=c).value = None

    # raw 값 채우기(세로)
    max_len = max((len(v) for v in hour_lists.values()), default=0)
    for row_i in range(max_len):
        r = RAW_START_ROW + row_i
        if r > RAW_END_ROW:
            break
        for i, h in enumerate(HOUR_ORDER):
            c = RAW_START_COL + i
            vals = hour_lists[h]
            ws.cell(row=r, column=c).value = vals[row_i] if row_i < len(vals) else None


def process_one_file(input_path: str, output_dir: str):
    xl = pd.ExcelFile(input_path)

    for lane_key, sheets in LANE_SHEETS.items():
        raw_lane = load_lane_raw(xl, sheets)

        for dtype, df_dtype in raw_lane.groupby("dtype"):
            if dtype == "UNKNOWN":
                continue

            template_path = TEMPLATES.get(dtype)
            if not template_path or not os.path.exists(template_path):
                raise FileNotFoundError(f"템플릿 파일을 찾을 수 없습니다: dtype={dtype}, path={template_path}")

            template_wb = openpyxl.load_workbook(template_path)
            out_wb = openpyxl.load_workbook(template_path)

            update_summary_lane_title_from_template(out_wb, template_wb, lane_key)

            for item_name, df_item in df_dtype.groupby("항목명"):
                sheet_name = item_to_sheetname(item_name)

                if sheet_name not in out_wb.sheetnames:
                    raise KeyError(
                        f"템플릿에 필요한 시트가 없습니다.\n"
                        f"dtype={dtype}, lane={lane_key}\n"
                        f"원본 항목명={item_name}\n"
                        f"찾는 시트명={sheet_name}\n"
                        f"템플릿 시트목록={out_wb.sheetnames}"
                    )

                ws = out_wb[sheet_name]
                hour_lists, mins, maxs, avgs = compute_hour_lists(df_item)

                fill_data_into_ws(
                    ws=ws,
                    dtype=dtype,
                    sheet_name=sheet_name,
                    hour_lists=hour_lists,
                    mins=mins,
                    maxs=maxs,
                    avgs=avgs,
                )

            out_path = os.path.join(output_dir, f"SLB_MES_{dtype}_Result_{lane_key}.xlsx")
            out_wb.save(out_path)

            out_wb.close()
            template_wb.close()


def main():
    import glob

    files = sorted(glob.glob(INPUT_GLOB))
    if not files:
        print(f"[WARN] 입력 파일이 없습니다. 패턴: {INPUT_GLOB}")
        return

    output_dir = "output"
    os.makedirs(output_dir, exist_ok=True)

    print(f"[INFO] 입력 파일 {len(files)}개 처리 시작...")
    for fp in files:
        try:
            print(f" - {fp}")
            process_one_file(fp, output_dir)
        except Exception as e:
            print(f"[ERROR] {fp} 처리 실패: {e}")

    print("[DONE] 완료!")


if __name__ == "__main__":
    main()
