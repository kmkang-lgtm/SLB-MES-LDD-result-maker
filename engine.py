# engine.py
# 원본 MES 엑셀(KHD/WPH) → 템플릿 기반 Result 엑셀(Lane1/Lane2) 생성 엔진
# - errors.py의 UserFacingError를 사용해 “사람이 읽기 쉬운” 에러를 던지도록 개선

import os
import re
import numpy as np
import pandas as pd
import openpyxl

from errors import (
    UserFacingError,
    missing_sheet_error,
    missing_column_error,
)

HOUR_ORDER = list(range(8, 24)) + [0] + list(range(1, 8))
HOUR_LABELS = list(range(8, 24)) + [24] + list(range(1, 8))

HEADER_ROW = 3

LANE_SHEETS = {
    "1Lane": ["1Lane_frt", "1Lane_rr", "1Lane_frt side", "1Lane_rr side"],
    "2Lane": ["2Lane_frt", "2Lane_rr", "2Lane_frt side", "2Lane_rr side"],
}


def detect_dtype(item_name: str) -> str:
    if "KHD" in item_name:
        return "KHD"
    if "WPH" in item_name:
        return "WPH"
    return "UNKNOWN"


def item_to_sheetname(item_name: str) -> str:
    name = item_name.replace("버스바 ", "")
    name = name.replace(" KHD AVG ", "-").replace(" WPH AVG ", "-")
    name = name.replace("FRT Side", "FS 1").replace("RR Side", "RS 1")
    return name


def safe_to_datetime(col):
    """
    문자열 파싱 실패가 많으면(>50%) 엑셀 날짜 일련번호로 재해석하는 안전장치
    """
    dt = pd.to_datetime(col, errors="coerce")
    if dt.isna().mean() > 0.5:
        num = pd.to_numeric(col, errors="coerce")
        dt2 = pd.to_datetime(num, unit="d", origin="1899-12-30", errors="coerce")
        if dt2.notna().sum() > dt.notna().sum():
            dt = dt2
    return dt


def _validate_required_columns(df: pd.DataFrame, required: list[str], *, where: str):
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise missing_column_error(
            needed=missing,
            available=[str(c) for c in df.columns.tolist()],
            title="원본 시트에 필요한 컬럼이 없습니다.",
            hint="원본 엑셀 포맷(컬럼명)이 변경됐는지 확인하세요.",
            context={"where": where},
        )


def load_lane_raw(xl: pd.ExcelFile, sheets, *, input_path: str, lane_key: str):
    """
    Lane별로 필요한 시트들을 읽어서 하나의 raw로 합침
    """
    # 시트 존재 검증(해석 쉬운 에러)
    available = xl.sheet_names
    for sh in sheets:
        if sh not in available:
            raise missing_sheet_error(
                needed=sh,
                available=available,
                title="원본 엑셀에 필요한 시트가 없습니다.",
                hint=f"파일({os.path.basename(input_path)})이 {lane_key} 포맷인지 확인하거나, 시트명이 변경됐는지 확인하세요.",
                context={"file": os.path.basename(input_path), "lane": lane_key},
            )

    dfs = []
    for sh in sheets:
        df = xl.parse(sh)
        _validate_required_columns(df, ["항목명", "측정일시", "측정값"], where=f"{lane_key}:{sh}")
        dfs.append(df)

    raw = pd.concat(dfs, ignore_index=True)

    raw["항목명"] = raw["항목명"].astype(str)
    raw["dtype"] = raw["항목명"].apply(detect_dtype)

    raw["측정일시"] = safe_to_datetime(raw["측정일시"])
    raw = raw.dropna(subset=["측정일시"])

    raw["hour"] = raw["측정일시"].dt.hour
    raw["val"] = pd.to_numeric(raw["측정값"], errors="coerce")
    return raw


def build_dynamic_hours(df_lane: pd.DataFrame, selected_hours=None):
    """
    Lane raw에서 실제로 존재하는 hour를 기반으로 동적 시간 컬럼을 생성
    - selected_hours가 있으면 그 시간만 남김
    - 템플릿 컬럼 수(기본 24) 초과 방지
    - 0시는 라벨을 24로 표시
    """
    hours = sorted(df_lane["hour"].dropna().unique().tolist())
    if not hours:
        hours = HOUR_ORDER[:]

    if selected_hours:
        hours = [h for h in hours if h in selected_hours]

    hours = hours[: len(HOUR_ORDER)]
    labels = [24 if h == 0 else h for h in hours]
    return hours, labels


def compute_hour_lists(df_item: pd.DataFrame, hour_order):
    hour_lists = {
        h: df_item.loc[df_item["hour"] == h, "val"].dropna().tolist()
        for h in hour_order
    }
    mins = [min(hour_lists[h]) if hour_lists[h] else 0 for h in hour_order]
    maxs = [max(hour_lists[h]) if hour_lists[h] else 0 for h in hour_order]
    avgs = [
        (sum(hour_lists[h]) / len(hour_lists[h])) if hour_lists[h] else np.nan
        for h in hour_order
    ]
    return hour_lists, mins, maxs, avgs


def update_summary_lane_title_from_template(out_wb, template_wb, lane_key):
    """
    템플릿 summary!B2 문자열의 맨 앞 숫자(1/2)만 lane에 맞게 교체
    """
    if "summary" not in out_wb.sheetnames or "summary" not in template_wb.sheetnames:
        # summary 탭이 없는 템플릿도 있을 수 있으니 조용히 패스
        return

    tpl_b2 = template_wb["summary"]["B2"].value or ""
    lane_no = "1" if lane_key.startswith("1") else "2"
    new_b2 = re.sub(r"^[12]", lane_no, str(tpl_b2))
    out_wb["summary"]["B2"].value = new_b2


def fill_data_into_ws(
    ws,
    dtype,
    sheet_name,
    hour_order,
    hour_labels,
    hour_lists,
    mins,
    maxs,
    avgs,
    raw_start_row=7,
    raw_end_row=100,
    raw_start_col=2,
    header_row=HEADER_ROW,
):
    """
    템플릿 서식/차트는 유지하고 값만 채움
    """
    max_cols = len(HOUR_ORDER)

    ws.cell(row=2, column=2).value = f"{dtype} {sheet_name}"

    # header row (시간 라벨)
    for i in range(len(hour_order)):
        c = raw_start_col + i
        ws.cell(row=header_row, column=c).value = hour_labels[i]

    for i in range(len(hour_order), max_cols):
        c = raw_start_col + i
        ws.cell(row=header_row, column=c).value = None

    # min/max/avg rows
    for i in range(len(hour_order)):
        c = raw_start_col + i
        ws.cell(row=4, column=c).value = mins[i]
        ws.cell(row=5, column=c).value = maxs[i]
        ws.cell(row=6, column=c).value = avgs[i]

    for i in range(len(hour_order), max_cols):
        c = raw_start_col + i
        ws.cell(row=4, column=c).value = None
        ws.cell(row=5, column=c).value = None
        ws.cell(row=6, column=c).value = None

    # raw 영역 클리어
    for r in range(raw_start_row, raw_end_row + 1):
        for i in range(max_cols):
            c = raw_start_col + i
            ws.cell(row=r, column=c).value = None

    # raw 값 채우기(세로)
    max_len = max((len(v) for v in hour_lists.values()), default=0)
    for row_i in range(max_len):
        r = raw_start_row + row_i
        if r > raw_end_row:
            break
        for i, h in enumerate(hour_order):
            c = raw_start_col + i
            vals = hour_lists[h]
            ws.cell(row=r, column=c).value = vals[row_i] if row_i < len(vals) else None


def make_results_for_input(
    input_path: str,
    templates: dict,
    output_dir: str,
    raw_end_row: int = 100,
    selected_hours=None,  # ✅ app.py에서 넘어오는 선택 시간대(list[int])
):
    """
    input_path: 원본 MES 엑셀 경로
    templates: {"KHD": ".../TEMPLATE_KHD.xlsx", "WPH": ".../TEMPLATE_WPH.xlsx"}
    output_dir: 결과 저장 폴더
    """
    if not os.path.exists(input_path):
        raise UserFacingError(
            title="입력 파일을 찾을 수 없습니다.",
            detail=f"경로: {input_path}",
            hint="업로드/경로를 다시 확인하세요.",
            code="INPUT_NOT_FOUND",
        )

    os.makedirs(output_dir, exist_ok=True)

    created = []
    try:
        xl = pd.ExcelFile(input_path)
    except Exception as e:
        raise UserFacingError(
            title="엑셀 파일을 열 수 없습니다.",
            detail=f"파일: {os.path.basename(input_path)}\n오류: {e}",
            hint="파일이 손상되었거나 암호/권한 문제가 있는지 확인하세요.",
            code="EXCEL_OPEN_FAILED",
        )

    for lane_key, sheets in LANE_SHEETS.items():
        raw_lane = load_lane_raw(xl, sheets, input_path=input_path, lane_key=lane_key)

        # 선택 시간 적용 후 시간이 하나도 없으면 사용자 친화 에러
        hour_order, hour_labels = build_dynamic_hours(raw_lane, selected_hours=selected_hours)
        if selected_hours and len(hour_order) == 0:
            raise UserFacingError(
                title="선택한 시간대에 데이터가 없습니다.",
                detail=f"선택 시간: {selected_hours}\n파일: {os.path.basename(input_path)} / {lane_key}",
                hint="다른 시간대를 선택하거나 원본 데이터 시간을 확인하세요.",
                code="NO_DATA_FOR_SELECTED_HOURS",
                context={"file": os.path.basename(input_path), "lane": lane_key},
            )

        for dtype, df_dtype in raw_lane.groupby("dtype"):
            if dtype == "UNKNOWN":
                continue

            if dtype not in templates:
                raise UserFacingError(
                    title="결과 생성에 필요한 템플릿이 없습니다.",
                    detail=f"dtype: {dtype}\ntemplates keys: {list(templates.keys())}",
                    hint="KHD/WPH 템플릿 경로가 올바른지 확인하세요.",
                    code="MISSING_TEMPLATE",
                )

            template_path = templates[dtype]
            if not os.path.exists(template_path):
                raise UserFacingError(
                    title="템플릿 파일을 찾을 수 없습니다.",
                    detail=f"dtype: {dtype}\n경로: {template_path}",
                    hint="템플릿 업로드/경로를 다시 확인하세요.",
                    code="TEMPLATE_NOT_FOUND",
                )

            template_wb = None
            out_wb = None
            try:
                template_wb = openpyxl.load_workbook(template_path)
                out_wb = openpyxl.load_workbook(template_path)

                update_summary_lane_title_from_template(out_wb, template_wb, lane_key)

                for item_name, df_item in df_dtype.groupby("항목명"):
                    sheet_name = item_to_sheetname(item_name)

                    if sheet_name not in out_wb.sheetnames:
                        raise UserFacingError(
                            title="템플릿에 필요한 탭(시트)이 없습니다.",
                            detail=(
                                f"[{dtype}] 찾는 탭: {sheet_name}\n"
                                f"원본 항목명: {item_name}\n"
                                f"템플릿 탭 목록: {out_wb.sheetnames}"
                            ),
                            hint="템플릿 탭 이름이 원본 '항목명' 변환 규칙과 일치하는지 확인하세요.",
                            code="MISSING_TEMPLATE_SHEET",
                            context={
                                "dtype": dtype,
                                "lane": lane_key,
                                "item_name": item_name,
                                "sheet_name": sheet_name,
                                "template": os.path.basename(template_path),
                            },
                        )

                    ws = out_wb[sheet_name]
                    hour_lists, mins, maxs, avgs = compute_hour_lists(df_item, hour_order)

                    fill_data_into_ws(
                        ws,
                        dtype,
                        sheet_name,
                        hour_order,
                        hour_labels,
                        hour_lists,
                        mins,
                        maxs,
                        avgs,
                        raw_end_row=raw_end_row,
                    )

                out_path = os.path.join(output_dir, f"SLB_MES_{dtype}_Result_{lane_key}.xlsx")
                out_wb.save(out_path)
                created.append(out_path)

            finally:
                # 워크북 닫기(예외가 나도 누수 방지)
                try:
                    if out_wb is not None:
                        out_wb.close()
                except Exception:
                    pass
                try:
                    if template_wb is not None:
                        template_wb.close()
                except Exception:
                    pass

    return created
