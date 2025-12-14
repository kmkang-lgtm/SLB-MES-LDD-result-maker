# dashboard_engine.py
# Summary 여러 개 → 기간 Dashboard 엑셀 생성 엔진
# - errors.py(UserFacingError) 기반으로 “사람이 읽기 쉬운” 에러를 던지도록 개선
# - 입력: zip bytes 또는 [(filename, bytes), ...]
# - 출력: (dashboard_filename, dashboard_bytes)

from __future__ import annotations

import io
import os
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Tuple, Optional

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

from errors import UserFacingError, invalid_zip_error


# ----------------------------
# Date helpers
# ----------------------------
def _parse_mmdd_from_name(name: str) -> Optional[str]:
    """
    파일명 끝에서 _(MM.DD).xlsx 또는 _MM.DD.xlsx 패턴을 찾으면 'MM.DD'를 반환
    """
    m = re.search(r"_\(?(\d{2}\.\d{2})\)?\.xlsx$", name)
    if m:
        return m.group(1)
    m2 = re.search(r"_(\d{2}\.\d{2})\.xlsx$", name)
    if m2:
        return m2.group(1)
    return None


def _date_key_from_filename_or_mtime(filename: str, mtime_ts: Optional[float] = None) -> str:
    """
    Summary 파일의 날짜키를 결정:
    1) 파일명에 _MM.DD가 있으면 그걸 사용
    2) 없으면 mtime 기반으로 MM.DD
    """
    mmdd = _parse_mmdd_from_name(filename)
    if mmdd:
        return mmdd

    if mtime_ts is None:
        # 최후 fallback: 현재일
        return datetime.now().strftime("%m.%d")
    return datetime.fromtimestamp(mtime_ts).strftime("%m.%d")


def _dashboard_name_from_dates(mmdd_list: List[str]) -> str:
    if not mmdd_list:
        return f"SLB_MES_Dashboard_{datetime.now().strftime('%m.%d')}.xlsx"
    # 정렬을 위해 MM.DD → (MM,DD)로 변환
    def _k(x):
        mm, dd = x.split(".")
        return int(mm), int(dd)

    s = sorted(mmdd_list, key=_k)
    return f"SLB_MES_Dashboard_{s[0]}~{s[-1]}.xlsx"


# ----------------------------
# Parsing Summary sheet
# ----------------------------
@dataclass
class BlockSpec:
    dtype: str           # "WPH" or "KHD"
    lane: str            # "1Lane" or "2Lane"
    start_col: int       # 1-based excel col index
    end_col: int         # inclusive


def _find_block_columns(header_row: List, token: str) -> List[int]:
    """
    header_row에서 'WPH'/'KHD' 토큰이 들어간 셀의 column index 리스트(0-based)를 리턴
    """
    cols = []
    for i, v in enumerate(header_row):
        if v is None:
            continue
        if token in str(v):
            cols.append(i)
    return cols


def _infer_blocks_from_sheet(ws) -> List[BlockSpec]:
    """
    Summary 포맷에서 상단 라벨을 기준으로 4개 블록(WPH/KHD x Lane1/2)을 찾는다.
    - 기존 코드 철학 유지: 'WPH' 포함 셀, 'KHD' 포함 셀의 시작 위치를 잡고,
      Lane1 블록 / Lane2 블록으로 분할.
    """
    # 보통 Summary는 1행 또는 2행에 라벨이 있는데, 안전하게 1~6행 탐색
    header_row_idx = None
    header_values = None
    for r in range(1, 7):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None and ("WPH" in str(v) or "KHD" in str(v)) for v in vals):
            header_row_idx = r
            header_values = vals
            break

    if header_row_idx is None:
        raise UserFacingError(
            title="Summary에서 WPH/KHD 헤더를 찾지 못했습니다.",
            detail="상단(1~6행)에서 'WPH' 또는 'KHD' 문자열이 포함된 셀을 찾을 수 없습니다.",
            hint="Summary 파일 형식이 변경되었는지 확인하세요.",
            code="SUMMARY_HEADER_NOT_FOUND",
        )

    wph_cols = _find_block_columns(header_values, "WPH")
    khd_cols = _find_block_columns(header_values, "KHD")

    if not wph_cols or not khd_cols:
        raise UserFacingError(
            title="Summary에서 WPH 또는 KHD 블록을 찾지 못했습니다.",
            detail=f"WPH cols: {wph_cols}\nKHD cols: {khd_cols}\nHeader row: {header_row_idx}",
            hint="Summary 파일 형식(라벨)이 바뀌었는지 확인하세요.",
            code="SUMMARY_BLOCK_NOT_FOUND",
        )

    # 보통 WPH 1Lane, KHD 1Lane, (공백열), WPH 2Lane, KHD 2Lane 순서
    # 라벨이 여러 열에 걸쳐 있을 수 있어 가장 왼쪽 시작을 기준으로 잡는다.
    wph_start = min(wph_cols) + 1
    khd_start = min(khd_cols) + 1

    # Lane2 쪽은 뒤쪽에 또 WPH/KHD가 반복되는 형태일 수 있음
    # 간단하게: 전체에서 WPH/KHD 위치를 기준으로 왼쪽 절반/오른쪽 절반을 나눔
    max_col = ws.max_column
    mid = max_col // 2

    wph_left = [c + 1 for c in wph_cols if (c + 1) <= mid]
    khd_left = [c + 1 for c in khd_cols if (c + 1) <= mid]
    wph_right = [c + 1 for c in wph_cols if (c + 1) > mid]
    khd_right = [c + 1 for c in khd_cols if (c + 1) > mid]

    if not wph_left or not khd_left or not wph_right or not khd_right:
        # 포맷이 mid 기준으로 안 나뉘는 경우가 있어 fallback: 가장 왼쪽 2개, 가장 오른쪽 2개로 잡기
        all_wph = sorted({c + 1 for c in wph_cols})
        all_khd = sorted({c + 1 for c in khd_cols})
        if len(all_wph) < 2 or len(all_khd) < 2:
            raise UserFacingError(
                title="Summary 블록을 Lane1/Lane2로 분리하지 못했습니다.",
                detail=f"WPH positions: {all_wph}\nKHD positions: {all_khd}",
                hint="Summary 시트의 레이아웃이 기존과 다른지 확인하세요.",
                code="SUMMARY_LANE_SPLIT_FAILED",
            )
        wph_left, wph_right = [all_wph[0]], [all_wph[-1]]
        khd_left, khd_right = [all_khd[0]], [all_khd[-1]]

    # 블록 범위를 정하려면 시작 col부터, 다음 블록 시작 전까지를 end로 잡는다.
    # 여기서는 각 시작 col을 기준으로 '연속된 유효 열'을 찾는 대신,
    # 통계 컬럼(Mean/Std/Min/Max/Range)이 대체로 5열 단위라는 점을 이용해 5열로 고정(기존 설계와 유사).
    # 필요시 5 → 6 등으로 조정 가능
    WIDTH = 5

    def _block(start_col: int, dtype: str, lane: str) -> BlockSpec:
        return BlockSpec(dtype=dtype, lane=lane, start_col=start_col, end_col=start_col + WIDTH - 1)

    blocks = [
        _block(min(wph_left), "WPH", "1Lane"),
        _block(min(khd_left), "KHD", "1Lane"),
        _block(min(wph_right), "WPH", "2Lane"),
        _block(min(khd_right), "KHD", "2Lane"),
    ]
    return blocks


def _read_block_to_tidy(ws, block: BlockSpec, mmdd: str) -> pd.DataFrame:
    """
    Summary 시트에서 블록 범위를 읽어 tidy 형태로 변환
    columns: date, dtype, lane, position, material, metric(Mean/Std/Min/Max/Range), value
    """
    start_col = block.start_col
    end_col = block.end_col

    # 헤더행을 다시 찾는다(블록 추정에 사용한 상단 라벨행과 별개로, 실제 표 헤더는 그 아래일 수 있음)
    # 안전하게 1~12행에서 "Position" 또는 "Material" 혹은 "Mean" 같은 단어를 찾는다.
    header_row = None
    for r in range(1, 13):
        row_vals = [ws.cell(row=r, column=c).value for c in range(start_col, end_col + 1)]
        joined = " ".join("" if v is None else str(v) for v in row_vals)
        if ("Mean" in joined) or ("Std" in joined) or ("Range" in joined) or ("Material" in joined) or ("Position" in joined):
            header_row = r
            break

    if header_row is None:
        # fallback: 1행을 헤더로 가정
        header_row = 1

    headers = [ws.cell(row=header_row, column=c).value for c in range(start_col, end_col + 1)]
    headers = [("" if h is None else str(h)).strip() for h in headers]

    # 첫 컬럼은 Position, 둘째는 Material일 수 있고, 나머지는 metrics일 수 있음.
    # 다양한 템플릿을 허용하기 위해, metrics는 헤더 문자열로 판단.
    # Position 컬럼 찾기
    pos_idx = None
    mat_idx = None
    metric_cols = []  # (metric_name, col_index)

    for i, h in enumerate(headers):
        hh = h.lower()
        if pos_idx is None and ("position" in hh or "pos" == hh):
            pos_idx = i
        if mat_idx is None and ("material" in hh or "mat" == hh):
            mat_idx = i

    # Position/Material 헤더가 없을 수도 있어 관습적 위치 fallback
    if pos_idx is None:
        pos_idx = 0
    if mat_idx is None and len(headers) >= 2:
        mat_idx = 1

    # metrics 후보: Mean/Std/Min/Max/Range가 포함된 헤더
    for i, h in enumerate(headers):
        hh = h.lower()
        if any(k in hh for k in ["mean", "avg", "std", "min", "max", "range"]):
            metric_cols.append((h if h else f"metric_{i}", i))

    if not metric_cols:
        # 마지막 3~5열을 metric으로 가정(최후 fallback)
        for i in range(max(0, len(headers) - 5), len(headers)):
            if i not in [pos_idx, mat_idx]:
                metric_cols.append((headers[i] if headers[i] else f"metric_{i}", i))

    # 데이터 시작 행: header_row + 1부터, position이 비면 종료
    records = []
    for r in range(header_row + 1, ws.max_row + 1):
        pos = ws.cell(row=r, column=start_col + pos_idx).value
        if pos is None or str(pos).strip() == "":
            continue  # 중간 빈행 허용
        pos_s = str(pos).strip()

        mat = None
        if mat_idx is not None:
            mat = ws.cell(row=r, column=start_col + mat_idx).value
        mat_s = "" if mat is None else str(mat).strip()

        # Material 규칙(요청사항: 현재 dashboard는 AL1/AL2를 쓰는 편)
        material_norm = _normalize_material_from_position(pos_s, mat_s)

        for metric_name, i in metric_cols:
            v = ws.cell(row=r, column=start_col + i).value
            if v is None or (isinstance(v, str) and v.strip() == ""):
                continue
            try:
                v_num = float(v)
            except Exception:
                continue

            records.append(
                {
                    "date": mmdd,
                    "dtype": block.dtype,
                    "lane": block.lane,
                    "position": pos_s,
                    "material": material_norm,
                    "metric": metric_name.strip() if metric_name else "metric",
                    "value": v_num,
                }
            )

    if not records:
        raise UserFacingError(
            title="Summary에서 유효한 데이터가 없습니다.",
            detail=f"block={block}\nmmdd={mmdd}\n(헤더행={header_row}, cols={start_col}~{end_col})",
            hint="Summary 파일이 비어있거나 레이아웃이 변경되었는지 확인하세요.",
            code="SUMMARY_NO_DATA",
        )

    return pd.DataFrame(records)


def _normalize_material_from_position(position: str, material_cell: str) -> str:
    """
    기존 dashboard_engine의 의도를 살려:
    - position에 '-2'가 있으면 AL 계열로 간주
      - '-1'이면 AL1, '-2'이면 AL2
    - 아니면 CU
    - material_cell이 이미 채워져 있으면 우선 사용(가능한 경우)
    """
    if material_cell:
        # 사용자가 이미 "AL1/AL2/CU/AL" 등을 넣어둔 경우 존중
        return material_cell

    # position suffix 판정
    m = re.search(r"-(\d+)\s*$", position)
    if m:
        suf = m.group(1)
        if suf == "1":
            return "AL1"
        if suf == "2":
            return "AL2"

    # '-2' 존재 여부 기반(원본 규칙 호환)
    if "-2" in position:
        return "AL2"
    return "CU"


# ----------------------------
# Build dashboard workbook
# ----------------------------
# ----------------------------
# Parsing (NEW Summary layout fallback)
# ----------------------------
def _cell_str(x) -> str:
    return str(x).strip() if x is not None else ""


def _is_new_summary_header_row(row_vals) -> bool:
    # expects: Lane | Position | Material | ...
    s = [_cell_str(v).lower() for v in row_vals[:3]]
    return len(s) >= 3 and s[0] == "lane" and s[1] == "position" and s[2] == "material"


def _normalize_lane(lane_cell: str) -> str:
    s = (lane_cell or "").strip().lower()
    # handle "Lane1", "1Lane", "1", etc.
    if "2" in s:
        return "2Lane"
    if "1" in s:
        return "1Lane"
    return ""


def _read_new_summary_sheet_to_tidy(ws, mmdd: str) -> pd.DataFrame:
    """
    Fallback parser for the new Summary format:
    - Table headers include: Lane, Position, Material, Mean_WPH, Mean_KHD, Std_WPH, ...
    - There may be two tables (Lane1 then Lane2) in the same sheet.
    Returns tidy df with columns:
      date, dtype, lane, position, material, metric, value
    """
    rows = []
    max_col = ws.max_column or 1
    for r in range(1, (ws.max_row or 1) + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        # trim trailing None
        while row and row[-1] is None:
            row.pop()
        rows.append(row)

    header_idxs = [i for i,row in enumerate(rows) if _is_new_summary_header_row(row)]
    if not header_idxs:
        raise UserFacingError(
            title="Summary 형식을 인식하지 못했습니다.",
            detail=f"파일의 Summary 시트에서 'Lane/Position/Material' 헤더를 찾지 못했습니다.",
            hint="Summary 파일 양식이 변경되었는지 확인하세요.",
            code="SUMMARY_NEW_FORMAT_HEADER_NOT_FOUND",
        )

    metric_pat = re.compile(r"^(Mean|Std|Min|Max|Range)_(WPH|KHD)$", re.IGNORECASE)

    records = []
    for hi, hidx in enumerate(header_idxs):
        headers = [str(x).strip() if x is not None else "" for x in rows[hidx]]
        # define table end: next header or end
        end = header_idxs[hi+1] if hi+1 < len(header_idxs) else len(rows)

        for rr in range(hidx + 1, end):
            row = rows[rr]
            if not row:
                continue

            # map row values to headers
            data = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            lane = _normalize_lane(_cell_str(data.get("Lane")))
            pos = _cell_str(data.get("Position"))
            if not pos:
                # blank position -> end of this table
                continue
            mat_cell = _cell_str(data.get("Material"))
            material = _normalize_material_from_position(pos, mat_cell)

            for col, val in data.items():
                m = metric_pat.match(col)
                if not m:
                    continue
                metric = m.group(1).capitalize()
                dtype = m.group(2).upper()

                if val is None or (isinstance(val, str) and val.strip() == ""):
                    continue
                try:
                    v = float(val)
                except Exception:
                    continue

                records.append(
                    {
                        "date": mmdd,
                        "dtype": dtype,
                        "lane": lane,
                        "position": pos,
                        "material": material,
                        "metric": metric,
                        "value": v,
                    }
                )

    if not records:
        raise UserFacingError(
            title="Summary에서 Dashboard용 숫자 데이터를 찾지 못했습니다.",
            detail="Mean_WPH/Mean_KHD 등의 숫자 컬럼을 파싱하지 못했습니다.",
            hint="Summary 파일이 비어있거나 컬럼명이 예상과 다른지 확인하세요.",
            code="SUMMARY_NEW_FORMAT_NO_DATA",
        )

    return pd.DataFrame(records)
def _write_dataframe(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1):
    # headers
    for j, col in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=j).value = col
    # body
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, v in enumerate(row, start=start_col):
            ws.cell(row=i, column=j).value = v


def _auto_fit_columns(ws, max_width: int = 50):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        maxlen = 0
        for r in range(1, min(ws.max_row, 2000) + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            s = str(v)
            if len(s) > maxlen:
                maxlen = len(s)
        ws.column_dimensions[letter].width = min(max(10, maxlen + 2), max_width)


def _style_as_table(ws, header_row: int = 1, freeze_panes: str = "A2", apply_filter: bool = True):
    """Lightweight styling to make sheets more readable."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F2F2F2")

    # header style
    for c in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = bold
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border

    # body borders
    for r in range(header_row + 1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = border

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze_panes

    if apply_filter and max_row >= header_row + 1:
        from openpyxl.utils import get_column_letter
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{max_row}"


def _add_heatmap_color_scale(ws, start_row: int = 2, start_col: int = 2):
    """Apply a color-scale conditional formatting to numeric area (simple heatmap)."""
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    if max_row < start_row or max_col < start_col:
        return

    rng = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(max_col)}{max_row}"
    rule = ColorScaleRule(start_type="min", start_color="FFF2F2F2",
                          mid_type="percentile", mid_value=50, mid_color="FFBFBFBF",
                          end_type="max", end_color="FF1F4E79")
    ws.conditional_formatting.add(rng, rule)
def _make_pivot(df_all: pd.DataFrame, metric_contains: str) -> pd.DataFrame:
    """
    df_all(tidy)에서 metric에 metric_contains가 포함된 값만 추출하여
    Pivot: index=[dtype,lane,material,position], columns=[date], values=value(mean) 형태로 만든다.
    """
    sub = df_all[df_all["metric"].str.contains(metric_contains, case=False, na=False)].copy()
    if sub.empty:
        raise UserFacingError(
            title=f"Pivot 생성 실패: '{metric_contains}' 데이터를 찾지 못했습니다.",
            detail=f"metric_contains={metric_contains}\n가능한 metric 예시: {df_all['metric'].dropna().unique().tolist()[:20]}",
            hint="Summary 파일의 metric 헤더(Mean/Std/Range 등)가 변경되었는지 확인하세요.",
            code="PIVOT_METRIC_NOT_FOUND",
        )

    pv = (
        sub.pivot_table(
            index=["dtype", "lane", "material", "position"],
            columns="date",
            values="value",
            aggfunc="mean",
        )
        .reset_index()
    )

    # columns 정렬: dtype,lane,material,position + 날짜들(정렬)
    date_cols = [c for c in pv.columns if c not in ["dtype", "lane", "material", "position"]]
    def _k(x):
        try:
            mm, dd = str(x).split(".")
            return int(mm), int(dd)
        except Exception:
            return (99, 99)
    date_cols = sorted(date_cols, key=_k)
    pv = pv[["dtype", "lane", "material", "position"] + date_cols]
    return pv


def _append_avg_min_max_formulas(ws, df_pivot: pd.DataFrame):
    """
    Pivot 시트의 날짜 컬럼들을 대상으로 마지막에 AVG/MIN/MAX 열을 추가하고 엑셀 수식 넣기
    """
    # 헤더는 1행, 데이터는 2행부터
    ncols = df_pivot.shape[1]
    # 날짜 시작 열 인덱스 찾기
    base_cols = 4  # dtype,lane,material,position
    date_start_col = base_cols + 1
    date_end_col = ncols

    avg_col = ncols + 1
    min_col = ncols + 2
    max_col = ncols + 3

    ws.cell(row=1, column=avg_col).value = "AVG"
    ws.cell(row=1, column=min_col).value = "MIN"
    ws.cell(row=1, column=max_col).value = "MAX"

    for r in range(2, ws.max_row + 1):
        start = f"{get_column_letter(date_start_col)}{r}"
        end = f"{get_column_letter(date_end_col)}{r}"
        rng = f"{start}:{end}"
        ws.cell(row=r, column=avg_col).value = f"=AVERAGE({rng})"
        ws.cell(row=r, column=min_col).value = f"=MIN({rng})"
        ws.cell(row=r, column=max_col).value = f"=MAX({rng})"


def build_dashboard_from_file_bytes(file_items: List[Tuple[str, bytes]]) -> Tuple[str, bytes]:
    """
    file_items: [(filename.xlsx, bytes), ...]
    return: (dashboard_filename, dashboard_bytes)
    """
    if not file_items:
        raise UserFacingError(
            title="Dashboard 생성에 필요한 Summary 파일이 없습니다.",
            hint="Summary 파일을 하나 이상 업로드하세요.",
            code="NO_SUMMARY_FILES",
        )

    # Summary 파일들을 읽어서 tidy로 합치기
    df_all = []
    mmdd_list = []

    for fname, fbytes in file_items:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(fbytes), data_only=False)
        except Exception as e:
            raise UserFacingError(
                title="Summary 엑셀을 열 수 없습니다.",
                detail=f"파일: {fname}\n오류: {e}",
                hint="파일 손상/권한/엑셀 형식 문제인지 확인하세요.",
                code="SUMMARY_OPEN_FAILED",
                context={"file": fname},
            )

        # 첫 시트 또는 활성 시트에 데이터가 있을 가능성이 큼
        ws = wb.active

        # 날짜키 결정
        mmdd = _date_key_from_filename_or_mtime(fname, None)
        mmdd_list.append(mmdd)

        try:
            blocks = _infer_blocks_from_sheet(ws)
            for b in blocks:
                tidy = _read_block_to_tidy(ws, b, mmdd)
                df_all.append(tidy)
        except Exception:
            # Fallback: new Summary layout (Lane/Position/Material + Mean_WPH/...)
            tidy = _read_new_summary_sheet_to_tidy(ws, mmdd)
            df_all.append(tidy)
        wb.close()

    df_all = pd.concat(df_all, ignore_index=True)

    # Dashboard 워크북 작성
    out_wb = openpyxl.Workbook()
    # 기본 시트 제거
    out_wb.remove(out_wb.active)

    ws_all = out_wb.create_sheet("All_Data")
    _write_dataframe(ws_all, df_all)
    _auto_fit_columns(ws_all)
    _style_as_table(ws_all, header_row=1, freeze_panes='A2', apply_filter=True)

    # Pivot sheets
    ws_pos = out_wb.create_sheet("ALL POSITION")
    # 포지션 목록(간단 요약)
    pos_df = (
        df_all[["dtype", "lane", "material", "position"]]
        .drop_duplicates()
        .sort_values(["dtype", "lane", "material", "position"])
        .reset_index(drop=True)
    )
    _write_dataframe(ws_pos, pos_df)
    _auto_fit_columns(ws_pos)
    _style_as_table(ws_pos, header_row=1, freeze_panes='A2', apply_filter=True)

    # Pivot Mean/Std/Range
    pv_mean = _make_pivot(df_all, "mean|avg")
    pv_std = _make_pivot(df_all, "std")
    pv_range = _make_pivot(df_all, "range")

    ws_pv_mean = out_wb.create_sheet("Pivot_Mean")
    _write_dataframe(ws_pv_mean, pv_mean)
    _append_avg_min_max_formulas(ws_pv_mean, pv_mean)
    _auto_fit_columns(ws_pv_mean)
    _style_as_table(ws_pv_mean, header_row=1, freeze_panes='A2', apply_filter=True)

    ws_pv_std = out_wb.create_sheet("Pivot_Std")
    _write_dataframe(ws_pv_std, pv_std)
    _append_avg_min_max_formulas(ws_pv_std, pv_std)
    _auto_fit_columns(ws_pv_std)
    _style_as_table(ws_pv_std, header_row=1, freeze_panes='A2', apply_filter=True)

    ws_pv_range = out_wb.create_sheet("Pivot_Range")
    _write_dataframe(ws_pv_range, pv_range)
    _append_avg_min_max_formulas(ws_pv_range, pv_range)
    _auto_fit_columns(ws_pv_range)
    _style_as_table(ws_pv_range, header_row=1, freeze_panes='A2', apply_filter=True)

    # Heatmap 시트는 기존 로직에 따라 더 꾸밀 수 있는데, 여기서는 기본 형태로 제공
    ws_hm = out_wb.create_sheet("Heatmap_Mean")
    _write_dataframe(ws_hm, pv_mean)  # 기본은 Pivot_Mean과 동일 데이터
    _auto_fit_columns(ws_hm)
    _style_as_table(ws_hm, header_row=1, freeze_panes='A2', apply_filter=True)
    _add_heatmap_color_scale(ws_hm, start_row=2, start_col=2)

    dash_name = _dashboard_name_from_dates(mmdd_list)

    out_buf = io.BytesIO()
    out_wb.save(out_buf)
    return dash_name, out_buf.getvalue()


def build_dashboard_from_zip_bytes(zip_bytes: bytes, zip_name: str) -> Tuple[str, bytes]:
    """
    zip_bytes: Summary 파일 여러 개가 들어있는 zip
    zip_name: 업로드된 zip 이름(표시용)
    """
    try:
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes), "r")
    except Exception as e:
        raise invalid_zip_error(
            detail=f"파일: {zip_name}\n오류: {e}",
            hint="올바른 ZIP 파일인지 확인하세요.",
            context={"zip": zip_name},
        )

    file_items: List[Tuple[str, bytes]] = []
    for n in zf.namelist():
        if n.lower().endswith(".xlsx") and not n.startswith("__MACOSX/"):
            try:
                file_items.append((os.path.basename(n), zf.read(n)))
            except Exception:
                continue

    if not file_items:
        raise UserFacingError(
            title="ZIP 안에서 Summary(.xlsx) 파일을 찾지 못했습니다.",
            detail=f"ZIP: {zip_name}\n내부 파일 목록: {zf.namelist()[:50]}",
            hint="Summary 엑셀들이 들어있는 ZIP을 업로드했는지 확인하세요.",
            code="NO_XLSX_IN_ZIP",
            context={"zip": zip_name},
        )

    return build_dashboard_from_file_bytes(file_items)